import requests
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout,
    QFileDialog, QMessageBox, QListWidget, QCheckBox
)
from PyQt5.QtGui import QPalette, QColor
from PyQt5.QtCore import Qt
from openpyxl import Workbook


class DHIS2DataSetRelatedDataelements(QWidget):
    def __init__(self):
        super().__init__()
        self.url_label = QLabel('URL:')
        self.url_input = QLineEdit()
        self.username_label = QLabel('Username:')
        self.username_input = QLineEdit()
        self.password_label = QLabel('Password:')
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.include_disaggregations = QCheckBox("Include Disaggregations")
        self.verify_button = QPushButton('Verify Credentials')
        self.dataset_list_label = QLabel('Available Data Sets:')
        self.dataset_list = QListWidget()
        self.dataset_list.setSelectionMode(QListWidget.MultiSelection)
        self.export_button = QPushButton('Export Data to XLS')
        self.export_button.setEnabled(False)

        self.init_ui()
        self.data_sets = {}

    def init_ui(self):
        layout = QVBoxLayout()
        layout.addWidget(self.url_label)
        layout.addWidget(self.url_input)
        layout.addWidget(self.username_label)
        layout.addWidget(self.username_input)
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(self.include_disaggregations)  # Add checkbox for disaggregations
        layout.addWidget(self.verify_button)
        layout.addWidget(self.dataset_list_label)
        layout.addWidget(self.dataset_list)
        layout.addWidget(self.export_button)

        self.setLayout(layout)
        self.setWindowTitle('DHIS2 Data Mapper')

        self.verify_button.clicked.connect(self.verify_credentials)
        self.export_button.clicked.connect(self.export_to_xls)

        # Apply theme
        self.apply_theme()

        self.show()

    def apply_theme(self):
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor("#eaf6ff"))  # Light blue background
        palette.setColor(QPalette.Base, QColor("#ffffff"))   # White text input background
        palette.setColor(QPalette.WindowText, QColor("#000000"))  # Black text
        self.setPalette(palette)

    def verify_credentials(self):
        base_url = self.url_input.text().strip()
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()

        if not base_url or not username or not password:
            QMessageBox.warning(self, "Input Error", "Please fill in all fields.")
            return

        auth = (username, password)
        data_sets_url = f"{base_url}/api/dataSets.json?fields=id,name&paging=false"

        response = requests.get(data_sets_url, auth=auth)
        if response.status_code == 200:
            data_sets = response.json().get('dataSets', [])
            self.data_sets = {ds['name']: ds['id'] for ds in data_sets}
            self.populate_dataset_list()
            QMessageBox.information(self, "Success", "Credentials verified and data sets fetched.")
        else:
            QMessageBox.critical(self, "Error", "Failed to verify credentials or fetch data sets.")

    def populate_dataset_list(self):
        self.dataset_list.clear()
        for name in self.data_sets.keys():
            self.dataset_list.addItem(name)
        self.export_button.setEnabled(True)

    def fetch_data_elements(self, base_url, username, password, dataset_id):
        auth = (username, password)

        # Adjust query based on checkbox state
        if self.include_disaggregations.isChecked():
            dataset_url = f"{base_url}/api/dataSets/{dataset_id}.json?fields=dataSetElements[dataElement[id,name,categoryCombo[categoryOptionCombos[id,name]]]]"
        else:
            dataset_url = f"{base_url}/api/dataSets/{dataset_id}.json?fields=dataSetElements[dataElement[id,name]]"

        response = requests.get(dataset_url, auth=auth)

        if response.status_code != 200:
            return None

        dataset_data = response.json()

        data_elements_info = []
        for dataSetElement in dataset_data.get('dataSetElements', []):
            data_element = dataSetElement['dataElement']
            data_element_id = data_element['id']
            data_element_name = data_element['name']

            if self.include_disaggregations.isChecked():
                # Include disaggregated data
                for combo in data_element['categoryCombo']['categoryOptionCombos']:
                    combo_id = combo['id']
                    combo_name = combo['name']
                    data_elements_info.append([data_element_id, data_element_name, combo_id, combo_name])
            else:
                # Only include default data elements
                data_elements_info.append([data_element_id, data_element_name])

        # Adjust columns based on checkbox state
        if self.include_disaggregations.isChecked():
            columns = ['Data Element ID', 'Data Element Name', 'Category Option Combo ID', 'Category Option Combo Name']
        else:
            columns = ['Data Element ID', 'Data Element Name']

        df = pd.DataFrame(data_elements_info, columns=columns)
        return df

    def export_to_xls(self):
        base_url = self.url_input.text().strip()
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()

        selected_items = self.dataset_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Selection Error", "Please select at least one data set.")
            return

        workbook = Workbook()
        for item in selected_items:
            dataset_name = item.text()
            dataset_id = self.data_sets[dataset_name]
            data_elements_df = self.fetch_data_elements(base_url, username, password, dataset_id)

            if data_elements_df is not None:
                sheet = workbook.create_sheet(title=dataset_name[:31])  # Sheet names max 31 chars
                for row in data_elements_df.itertuples(index=False, name=None):
                    sheet.append(row)

        # Remove the default empty sheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        # Save the workbook
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Workbook", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_path:
            workbook.save(file_path)
            QMessageBox.information(self, "Success", "Data exported successfully!")


if __name__ == '__main__':
    app = QApplication([])
    window = DHIS2DataSetRelatedDataelements()
    app.exec_()
